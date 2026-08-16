# Generate the blank annotator-2 IAA sheet for the gold-triple audit (R01 M3).
# Reproduces the EXACT seed-42 50-row sample per dataset (same as gold_triple_audit.py --sample 50),
# so ann-2's rows align by (idx, h, r, t) with the LLM verdicts + annotator-1 labels produced on 25/07.
import ast, os, random
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = r"E:/git-workspace/edc"
OUT = r"E:/git-workspace/edc/assets/review/IAA_annotator2_gold_audit.xlsx"

def load_kg_txt(path):
    data = []
    for line in open(path, encoding="utf-8"):
        line = line.strip()
        if not line:
            data.append([]); continue
        try: data.append(ast.literal_eval(line))
        except Exception: data.append([])
    return data

def load_src(path):
    return [l.rstrip("\n") for l in open(path, encoding="utf-8")]

def sample_ds(ds, k=50, seed=42):
    gold = load_kg_txt(f"{ROOT}/edca/evaluate/references/{ds}.txt")
    src = load_src(f"{ROOT}/datasets/{ds}.txt")
    n = min(len(gold), len(src))
    items = []
    for i in range(n):
        for t in (gold[i] or []):
            if isinstance(t, (list, tuple)) and len(t) >= 3:
                items.append({"idx": i, "h": str(t[0]), "r": str(t[1]), "t": str(t[2]), "sent": src[i]})
    if k < len(items):
        items = random.Random(seed).sample(items, k)
    return items

wb = openpyxl.Workbook()
# ---- Instructions tab ----
ins = wb.active; ins.title = "Instructions"
HEAD = Font(bold=True, size=13); B = Font(bold=True)
ins["A1"] = "Gold-triple audit — Annotator 2 (inter-annotator agreement)"; ins["A1"].font = HEAD
lines = [
 "",
 "Purpose: a second, independent human labels the SAME 50 gold triples/dataset that the LLM judge and",
 "Annotator 1 label, so we can report inter-annotator kappa (addresses reviewer concern on single-annotator gold).",
 "",
 "Task: for each row, decide whether the benchmark GOLD triple (head, relation, tail) is actually expressed",
 "by its source SENTENCE. Fill exactly three cells per row (q1, q2, q3). The verdict is derived automatically.",
 "",
 "q1_head_in_sentence : is the HEAD entity mentioned in the sentence (by name, alias, pronoun, or",
 "                      definite description)?  -> yes / no",
 "q2_tail_in_sentence : is the TAIL entity mentioned likewise?  -> yes / no",
 "q3_relation_expressed: is the RELATION between them expressed?  -> stated / inferable / no",
 "   - stated   : READABLE from the sentence surface via ANY construction (verb, copula, apposition,",
 "                preposition, parenthetical, genitive, or compound), needing NO outside knowledge.",
 "   - inferable: NOT surface-expressed, but strictly follows from the sentence content by genuine inference.",
 "   - no       : the sentence does not support this relation.",
 "",
 "Derived verdict (do NOT fill; for reference):",
 "   q1=no OR q2=no                 -> NOT_VERBALIZED  (distant-supervision artifact; not in the text)",
 "   q1=yes & q2=yes & q3=no        -> UNSUPPORTED     (label-error candidate)",
 "   q1=yes & q2=yes & q3=inferable -> IMPLICIT_OK     (true, inferred)",
 "   q1=yes & q2=yes & q3=stated    -> SUPPORTED       (true, surface-stated)",
 "",
 "Label independently — do not consult the LLM output or Annotator 1. Use the 'notes' column for hard cases.",
 "One tab per dataset: WebNLG, REBEL, Wiki-NRE (50 rows each).",
]
for i, t in enumerate(lines, start=2):
    ins[f"A{i}"] = t
ins.column_dimensions["A"].width = 110

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_fill = PatternFill("solid", fgColor="1F6CC4")
fill_col = PatternFill("solid", fgColor="FFF3D6")  # highlight the 3 fill columns
cols = ["idx", "sentence", "head", "relation", "tail",
        "q1_head_in_sentence", "q2_tail_in_sentence", "q3_relation_expressed", "notes"]
widths = [6, 70, 22, 26, 22, 20, 20, 22, 30]

for ds, tab in [("webnlg", "WebNLG"), ("rebel", "REBEL"), ("wiki-nre", "Wiki-NRE")]:
    rows = sample_ds(ds)
    ws = wb.create_sheet(tab)
    for c, (name, w) in enumerate(zip(cols, widths), start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = hdr_fill
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[chr(64+c)].width = w
    for r, it in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=it["idx"])
        ws.cell(row=r, column=2, value=it["sent"]).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=3, value=it["h"])
        ws.cell(row=r, column=4, value=it["r"])
        ws.cell(row=r, column=5, value=it["t"])
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = border
            if c in (6, 7, 8):
                ws.cell(row=r, column=c).fill = fill_col
    # dropdowns
    dv_yn = DataValidation(type="list", formula1='"yes,no"', allow_blank=True)
    dv_q3 = DataValidation(type="list", formula1='"stated,inferable,no"', allow_blank=True)
    ws.add_data_validation(dv_yn); ws.add_data_validation(dv_q3)
    last = len(rows) + 1
    dv_yn.add(f"F2:F{last}"); dv_yn.add(f"G2:G{last}"); dv_q3.add(f"H2:H{last}")
    ws.freeze_panes = "A2"

wb.save(OUT)
print("saved", OUT)
for ds in ["webnlg","rebel","wiki-nre"]:
    print(ds, "sample rows:", len(sample_ds(ds)))
