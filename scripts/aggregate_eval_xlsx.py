"""
Aggregate every per-run eval output into ONE human-viewable .xlsx (one sheet per
metric family), so the full matrix can be reviewed/verified at a glance and dropped
into the paper. Pure-python (pandas+openpyxl); no model, no GPU. (DECISIONS 2026-06-25.)

Discovers the 6 matrix cells (mode {1,2,3} x retrieval {item, itemcluster}) by the
run-folder naming convention, reads whatever eval_*/ files exist (robust to missing —
Mode 1 has no error_attribution/retrieval), and writes:

  Overview            one row per run + key OIE/size numbers (the table-of-contents)
  Triple_raw          9 cases x [6 cells] x {partial,strict,exact}  (RAW; the headline)
  Triple_normalized   same, literal-normalized (dates->ISO, numbers canonical)
  Literal_gap         per cell x metric: norm - raw delta (the "literal-format gap")
  Bcubed              9 cases x [6 cells] x {B3 F1/P/R, pairwise, #clusters}
  Entity_canon        entity-type canon ec1/2/3 x {B3,ARI,V,#predT,#goldT}  (all slots)
  Entity_grounded     same, scored only on KB-confirmed/accepted slots + grounded coverage
  Entity_p31          same, gold type = per-entity instance-of (identity), not slot role (ENTITY_TYPE_SCHEMA §10)
  Mode1_aligned       Mode-1 name-drift-robust aligned-F1 / P / R / #pred_rel
  Error_attribution   Mode 2/3: OIE/RETRIEVAL/CANON/CORRECT %
  Retrieval_recall    Mode 2/3: recall@1/3/5, gold-in-pool, mean rank
  OIE_diagnosis       per run: A/B/C/D/E/F buckets + 3-way rollup
  Intrinsic           type-conformance / components / isolated / reduction

Per-column winner is bold + green-filled.

Usage (standalone, e.g. webnlg already done):
    python scripts/aggregate_eval_xlsx.py --dataset webnlg --model qwen3-8b \
        --emb bgem3 --gpu A100 --date_tag 20260622 \
        --extra-runs output/webnlg_selfcanon2_mode3_item_qwen3-8b_bgem3_A100_20260622_refe2e1
Auto (patched into run_matrix.sh after the matrix completes).
"""

import os
import re
import csv
import json
import glob
import argparse

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

CASES = [
    "case1_embed_threshold", "case2_name_only", "case3_name_gendef_edc",
    "case4_name_gendef_abstract", "case5_name_detail", "case6_name_detail_headtail",
    "case7_detail_typed", "case8_concat", "case9_weighted",
]

_BOLD = Font(bold=True)
_CENTER = Alignment(horizontal="center")
_GREEN = PatternFill("solid", fgColor="C6EFCE")
_GREY = PatternFill("solid", fgColor="EDEDED")


# ---------------------------------------------------------------------------
# tiny readers (robust: missing file -> {})
# ---------------------------------------------------------------------------
def _f(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return x


def read_csv_by(path, key="case"):
    """CSV -> {key_value: {col: num_or_str}}. {} if missing."""
    if not os.path.isfile(path):
        return {}
    out = {}
    with open(path, newline="", encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            k = row.get(key)
            if k is None:
                continue
            out[k] = {c: _f(v) for c, v in row.items() if c != key}
    return out


def read_json(path):
    if not os.path.isfile(path):
        return {}
    with open(path, encoding="utf-8") as fh:
        return json.load(fh)


# ---------------------------------------------------------------------------
# run discovery
# ---------------------------------------------------------------------------
def discover(dataset, model, emb, gpu, date_tag, iterdir, extra):
    """Return ordered list of (label, mode, retr, iter_dir)."""
    runs = []
    for mode in (1, 2, 3):
        for retr, short in (("item", "item"), ("itemcluster", "cluster")):
            base = f"output/{dataset}_selfcanon2_mode{mode}_{retr}_{model}_{emb}_{gpu}_{date_tag}"
            d = os.path.join(base, iterdir)
            if os.path.isdir(d):
                runs.append((f"M{mode}_{short}", mode, retr, d))
    for e in extra or []:
        d = e if os.path.basename(e) == iterdir else os.path.join(e, iterdir)
        if not os.path.isdir(d):
            print(f"[warn] extra run not found: {d}")
            continue
        name = os.path.basename(os.path.dirname(d))
        m = re.search(r"mode(\d)_(item|itemcluster)", name)
        suffix = name.split(date_tag)[-1].lstrip("_") if date_tag in name else ""
        lbl = (f"M{m.group(1)}_{'cluster' if m.group(2)=='itemcluster' else 'item'}"
               + (f":{suffix}" if suffix else ":extra")) if m else name[:18]
        mode = int(m.group(1)) if m else 0
        retr = m.group(2) if m else "item"
        runs.append((lbl, mode, retr, d))
    return runs


# ---------------------------------------------------------------------------
# generic 2-tier grouped sheet with per-column winner highlight
# ---------------------------------------------------------------------------
def add_sheet(wb, name, row_keys, row_header, groups, best="max", numfmt="0.0000"):
    """groups = [(group_label_or_'', [(sub_label, {row_key: value}), ...]), ...]."""
    ws = wb.create_sheet(name[:31])
    ws.cell(row=2, column=1, value=row_header).font = _BOLD
    col = 2
    metas = []
    for glabel, subs in groups:
        start = col
        for sub_label, valmap in subs:
            h = ws.cell(row=2, column=col, value=sub_label)
            h.font = _BOLD
            h.alignment = _CENTER
            metas.append((col, valmap))
            col += 1
        end = col - 1
        if glabel:
            g = ws.cell(row=1, column=start, value=glabel)
            g.font = _BOLD
            g.alignment = _CENTER
            g.fill = _GREY
            if end >= start:
                ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
    for r, rk in enumerate(row_keys, start=3):
        ws.cell(row=r, column=1, value=rk)
        for col_idx, valmap in metas:
            v = valmap.get(rk)
            cell = ws.cell(row=r, column=col_idx, value=v)
            if isinstance(v, float):
                cell.number_format = numfmt
    if best:
        for col_idx, valmap in metas:
            nums = [valmap.get(rk) for rk in row_keys if isinstance(valmap.get(rk), (int, float))]
            if not nums:
                continue
            target = max(nums) if best == "max" else min(nums)
            for r, rk in enumerate(row_keys, start=3):
                if valmap.get(rk) == target:
                    c = ws.cell(row=r, column=col_idx)
                    c.font = _BOLD
                    c.fill = _GREEN
    ws.freeze_panes = "B3"
    ws.column_dimensions["A"].width = max(14, len(row_header) + 2,
                                          *(len(str(k)) + 1 for k in row_keys)) if row_keys else 18
    for ci in range(2, col):
        ws.column_dimensions[get_column_letter(ci)].width = 9.5
    return ws


# ---------------------------------------------------------------------------
# build
# ---------------------------------------------------------------------------
def build(runs, out_path):
    wb = Workbook()
    wb.remove(wb.active)

    # load everything once
    data = {}
    for lbl, mode, retr, d in runs:
        data[lbl] = {
            "mode": mode, "retr": retr, "dir": d,
            "t1": read_csv_by(os.path.join(d, "eval/table1_closed_schema.csv")),
            "t1n": read_csv_by(os.path.join(d, "eval/table1_closed_schema_normalized.csv")),
            "gap": read_csv_by(os.path.join(d, "eval/table1_literal_gap.csv")),
            "bc": read_csv_by(os.path.join(d, "eval_clustering/clustering_metrics.csv")),
            "ent": read_csv_by(os.path.join(d, "eval_clustering/clustering_entity_metrics.csv"), key="ec_case"),
            "entg": read_csv_by(os.path.join(d, "eval_clustering/clustering_entity_grounded.csv"), key="ec_case"),
            "entp31": read_csv_by(os.path.join(d, "eval_clustering/clustering_entity_p31.csv"), key="ec_case"),
            "m1": read_csv_by(_first(d, "eval_mode1", "*.csv")),
            "err": read_csv_by(os.path.join(d, "eval_error_attribution/error_attribution.csv")),
            "rec": read_csv_by(os.path.join(d, "eval_retrieval/retrieval_recall.csv")),
            "oie": read_json(os.path.join(d, "eval_oie/oie_metrics.json")),
            "miss": read_json(os.path.join(d, "eval_oie_miss/oie_miss_diagnosis.json")),
            "intr": read_csv_by(_first(d, "eval_intrinsic", "*.csv")),
        }
    labels = list(data)

    # ---- Overview ----
    ov_rows = []
    ov_cols = ["mode", "retr", "n_gold", "OIE_ent_F1", "OIE_triple_F1", "count_ratio",
               "OIE_MISS%", "genuine%(C+D)", "surface%(B+F)"]
    ov = {c: {} for c in ov_cols}
    for lbl in labels:
        dd = data[lbl]
        ov["mode"][lbl] = dd["mode"]
        ov["retr"][lbl] = dd["retr"]
        m = dd["miss"].get("overall", {})
        ov["n_gold"][lbl] = m.get("n_gold")
        ov["OIE_MISS%"][lbl] = m.get("MISS_%")
        ov["genuine%(C+D)"][lbl] = m.get("rollup_genuine_extraction_%")
        ov["surface%(B+F)"][lbl] = m.get("rollup_surface_or_format_%")
        o = dd["oie"]
        ov["OIE_ent_F1"][lbl] = o.get("entity_f1")
        ov["OIE_triple_F1"][lbl] = o.get("triple_soft_f1")
        ov["count_ratio"][lbl] = o.get("count_ratio_mean(pred/gold)")
    add_sheet(wb, "Overview", labels, "run", [("", [(c, ov[c]) for c in ov_cols])], best=None)

    # ---- Triple sheets (raw / normalized / gap) ----
    for sheet, key, fld_best in (("Triple_raw", "t1", "max"),
                                 ("Triple_normalized", "t1n", "max")):
        groups = []
        for lbl in labels:
            t = data[lbl][key]
            groups.append((lbl, [(m, {c: t.get(c, {}).get(m) for c in CASES})
                                 for m in ("partial", "strict", "exact")]))
        add_sheet(wb, sheet, CASES, "case", groups, best=fld_best)
    # literal gap (delta only; norm = raw + delta)
    groups = []
    for lbl in labels:
        g = data[lbl]["gap"]
        groups.append((lbl, [(m, {c: g.get(c, {}).get(m) for c in CASES})
                             for m in ("partial_gap", "strict_gap", "exact_gap")]))
    add_sheet(wb, "Literal_gap", CASES, "case", groups, best=None)

    # ---- B-cubed (+ chance-adjusted ARI/NMI/V) ----
    groups = []
    for lbl in labels:
        b = data[lbl]["bc"]
        groups.append((lbl, [(m, {c: b.get(c, {}).get(src) for c in CASES})
                             for m, src in (("B3_F1", "bcubed_f1"), ("ARI", "ari"), ("NMI", "nmi"),
                                            ("V", "v_measure"), ("#cl", "n_pred_clusters"))]))
    add_sheet(wb, "Bcubed", CASES, "case", groups)

    # ---- Entity-type canon clustering (rows = ec1/ec2/ec3) ----
    ent_labels = [l for l in labels if data[l]["ent"]]
    if ent_labels:
        ec_rows = sorted({k for l in ent_labels for k in data[l]["ent"]})
        groups = []
        for lbl in ent_labels:
            e = data[lbl]["ent"]
            groups.append((lbl, [(m, {ec: e.get(ec, {}).get(src) for ec in ec_rows})
                                 for m, src in (("B3_F1", "bcubed_f1"), ("ARI", "ari"),
                                                ("V", "v_measure"), ("#predT", "n_pred_clusters"),
                                                ("#goldT", "n_gold_clusters"))]))
        add_sheet(wb, "Entity_canon", ec_rows, "ec_case", groups)

    # ---- KB-grounded entity-type canon (scored only on KB-confirmed/accepted slots) ----
    entg_labels = [l for l in labels if data[l]["entg"]]
    if entg_labels:
        ec_rows = sorted({k for l in entg_labels for k in data[l]["entg"]})
        groups = []
        for lbl in entg_labels:
            e = data[lbl]["entg"]
            groups.append((lbl, [(m, {ec: e.get(ec, {}).get(src) for ec in ec_rows})
                                 for m, src in (("B3_F1", "bcubed_f1"), ("ARI", "ari"),
                                                ("V", "v_measure"), ("cov", "grounded_coverage"),
                                                ("#predT", "n_pred_clusters"), ("#goldT", "n_gold_clusters"))]))
        add_sheet(wb, "Entity_grounded", ec_rows, "ec_case", groups)

    # ---- per-entity P31 gold entity-type canon (identity-based reference; ENTITY_TYPE_SCHEMA §10) ----
    entp_labels = [l for l in labels if data[l]["entp31"]]
    if entp_labels:
        ec_rows = sorted({k for l in entp_labels for k in data[l]["entp31"]})
        groups = []
        for lbl in entp_labels:
            e = data[lbl]["entp31"]
            groups.append((lbl, [(m, {ec: e.get(ec, {}).get(src) for ec in ec_rows})
                                 for m, src in (("B3_F1", "bcubed_f1"), ("ARI", "ari"),
                                                ("V", "v_measure"), ("cov", "grounded_coverage"),
                                                ("#predT", "n_pred_clusters"), ("#goldT", "n_gold_clusters"))]))
        add_sheet(wb, "Entity_p31", ec_rows, "ec_case", groups)

    # ---- Mode1 aligned (mode-1 runs only) ----
    m1_labels = [l for l in labels if data[l]["mode"] == 1 and data[l]["m1"]]
    if m1_labels:
        groups = []
        for lbl in m1_labels:
            mm = data[lbl]["m1"]
            groups.append((lbl, [(m, {c: mm.get(c, {}).get(src) for c in CASES})
                                 for m, src in (("aligned_F1", "aligned_triple_f1"),
                                                ("P", "aligned_triple_precision"),
                                                ("R", "aligned_triple_recall"),
                                                ("#rel", "n_pred_relations"))]))
        add_sheet(wb, "Mode1_aligned", CASES, "case", groups)

    # ---- Error attribution (mode 2/3) ----
    ea_labels = [l for l in labels if data[l]["err"]]
    if ea_labels:
        groups = []
        for lbl in ea_labels:
            e = data[lbl]["err"]
            groups.append((lbl, [(m, {c: e.get(c, {}).get(src) for c in CASES})
                                 for m, src in (("CORRECT", "CORRECT_pct"), ("OIE", "OIE_MISS_pct"),
                                                ("RETR", "RETRIEVAL_MISS_pct"), ("CANON", "CANON_MISS_pct"))]))
        add_sheet(wb, "Error_attribution", CASES, "case", groups, best=None)

    # ---- Retrieval recall (mode 2/3) ----
    rc_labels = [l for l in labels if data[l]["rec"]]
    if rc_labels:
        groups = []
        for lbl in rc_labels:
            rr = data[lbl]["rec"]
            sub = [(k, src) for k, src in (("R@1", "recall@1"), ("R@3", "recall@3"),
                                           ("R@5", "recall@5"), ("pool", "gold_in_pool_rate"))]
            groups.append((lbl, [(m, {c: rr.get(c, {}).get(src) for c in CASES}) for m, src in sub]))
        add_sheet(wb, "Retrieval_recall", CASES, "case", groups)

    # ---- OIE diagnosis (per run overall) ----
    od_cols = [("CORRECT", "CORRECT_%"), ("B_surface", "B_surface_%"), ("F_literal", "F_literal_format_%"),
               ("C_entity", "C_entity_miss_%"), ("D_pairing", "D_pairing_miss_%"),
               ("A_unreal", "A_unrealized_%"), ("E_coref", "E_coref_%"),
               ("surf/fmt", "rollup_surface_or_format_%"), ("genuine", "rollup_genuine_extraction_%"),
               ("ceiling", "rollup_ceiling_or_linking_%")]
    od = {m: {} for m, _ in od_cols}
    for lbl in labels:
        o = data[lbl]["miss"].get("overall", {})
        for m, src in od_cols:
            od[m][lbl] = o.get(src)
    add_sheet(wb, "OIE_diagnosis", labels, "run", [("", [(m, od[m]) for m, _ in od_cols])], best=None)

    # ---- Intrinsic ----
    it_labels = [l for l in labels if data[l]["intr"]]
    if it_labels:
        # pick a few interpretable columns if present
        sample = next(iter(data[it_labels[0]]["intr"].values()))
        want = [c for c in ("type_conformance", "num_components", "num_isolated_nodes",
                            "relation_reduction_ratio", "relation_type_entropy") if c in sample]
        groups = []
        for lbl in it_labels:
            ii = data[lbl]["intr"]
            groups.append((lbl, [(c.split("_")[0][:6], {cs: ii.get(cs, {}).get(c) for cs in CASES})
                                 for c in want]))
        if want:
            add_sheet(wb, "Intrinsic", CASES, "case", groups, best=None)

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    print(f"Saved -> {out_path}  ({len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames)})")
    print(f"Runs aggregated: {', '.join(labels)}")


def _first(d, sub, pat):
    hits = glob.glob(os.path.join(d, sub, pat))
    return hits[0] if hits else os.path.join(d, sub, "_missing_")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dataset", required=True)
    ap.add_argument("--model", default="qwen3-8b")
    ap.add_argument("--emb", default="bgem3")
    ap.add_argument("--gpu", default="A100")
    ap.add_argument("--date_tag", required=True)
    ap.add_argument("--iter", default="iter0")
    ap.add_argument("--extra-runs", nargs="*", default=[])
    ap.add_argument("--out")
    args = ap.parse_args()

    runs = discover(args.dataset, args.model, args.emb, args.gpu, args.date_tag, args.iter, args.extra_runs)
    if not runs:
        raise SystemExit(f"[aggregate] no runs found for {args.dataset} {args.date_tag} "
                         f"({args.model}/{args.emb}/{args.gpu})")
    out = args.out or f"output/eval_summary_{args.dataset}_{args.date_tag}.xlsx"
    build(runs, out)


if __name__ == "__main__":
    main()
